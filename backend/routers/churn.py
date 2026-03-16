"""
/api/stores/churn  –  Store churn analysis
Butikker aktive i 2025 men ikke set i 2026 endnu.
Status: "Ophørt" = ikke i api_stores_list.xlsx / "Pause" = er i listen men ingen 2026-posts.
"""
from fastapi import APIRouter, Query
from pathlib import Path
from collections import defaultdict
from database import get_conn
import openpyxl

router = APIRouter(prefix="/api/stores/churn", tags=["churn"])

# api_stores_list.xlsx lives in backend/data/
_XLSX     = Path(__file__).parent.parent / "data" / "api_stores_list.xlsx"
_XLSX2026 = Path(__file__).parent.parent / "data" / "active_stores_list_2026.xlsx"

_active_kardex_cache: set | None = None
_active_2026_cache:   set | None = None
_store_name_map_cache: dict | None = None


def _get_active_kardex() -> set:
    """Return set of kardex_ids from the official active-store Excel list."""
    global _active_kardex_cache
    if _active_kardex_cache is None:
        ids: set = set()
        if _XLSX.exists():
            wb = openpyxl.load_workbook(_XLSX, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] is not None:
                    ids.add(int(row[1]))
            wb.close()
        _active_kardex_cache = ids
    return _active_kardex_cache


def _get_active_2026_kardex() -> set:
    """Return set of kardex_ids from active_stores_list_2026.xlsx (Navn + Kardex)."""
    global _active_2026_cache
    if _active_2026_cache is None:
        ids: set = set()
        if _XLSX2026.exists():
            wb = openpyxl.load_workbook(_XLSX2026, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] is not None:
                    ids.add(int(row[1]))
            wb.close()
        _active_2026_cache = ids
    return _active_2026_cache


def _get_store_name_map() -> dict:
    """Return {kardex_id: store_name} from both xlsx files (api_stores_list first, 2026 second)."""
    global _store_name_map_cache
    if _store_name_map_cache is None:
        names: dict = {}
        for xlsx_path in (_XLSX, _XLSX2026):
            if xlsx_path.exists():
                wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None and row[1] is not None:
                        kid = int(row[1])
                        if kid not in names:  # api_stores_list takes priority
                            names[kid] = str(row[0])
                wb.close()
        _store_name_map_cache = names
    return _store_name_map_cache


def _norm_chain(store_name: str) -> str:
    """Return short display name: 'Kvickly Hovedkontor' → 'Kvickly', etc."""
    for prefix in ("Kvickly", "Superbrugsen"):
        if store_name.startswith(prefix):
            return prefix
    return store_name


# Base CTE: inactive = active in 2025 AND NOT active in 2026
_CTE = """
WITH stats_2025 AS (
    SELECT
        kardex_id,
        ANY_VALUE(store_name)       AS store_name,
        COUNT(*)                    AS offer_count,
        ROUND(AVG(jatak_count), 1)  AS avg_jatak,
        MAX(created_date)           AS seneste_opslag
    FROM jatak
    WHERE YEAR(created_date) = 2025
    GROUP BY kardex_id
),
active_2026 AS (
    SELECT DISTINCT kardex_id FROM jatak WHERE YEAR(created_date) = 2026
),
inactive AS (
    SELECT s.* FROM stats_2025 s
    LEFT JOIN active_2026 a ON s.kardex_id = a.kardex_id
    WHERE a.kardex_id IS NULL
)
"""


@router.get("/summary")
def get_churn_summary():
    """
    KPI totals + chain-level breakdown for stores active in 2025 but not in 2026.
    """
    conn = get_conn()
    active_ids  = _get_active_kardex()
    active_2026 = _get_active_2026_kardex()

    rows = conn.execute(_CTE + """
        SELECT kardex_id, store_name, offer_count, avg_jatak, seneste_opslag
        FROM inactive
        ORDER BY offer_count DESC
    """).fetchall()

    total_active_2025 = int(conn.execute(
        "SELECT COUNT(DISTINCT kardex_id) FROM jatak WHERE YEAR(created_date) = 2025"
    ).fetchone()[0])

    total_offers_2025 = int(conn.execute(
        "SELECT COUNT(*) FROM jatak WHERE YEAR(created_date) = 2025"
    ).fetchone()[0])

    new_stores_2026 = int(conn.execute("""
        SELECT COUNT(DISTINCT kardex_id) FROM jatak
        WHERE YEAR(created_date) = 2026
          AND kardex_id NOT IN (
              SELECT DISTINCT kardex_id FROM jatak WHERE YEAR(created_date) = 2025
          )
    """).fetchone()[0])

    # Enrich each store with chain + status + hk_opslag
    stores = []
    for r in rows:
        kid    = int(r[0])
        is_pause = kid in active_ids
        stores.append({
            "kardex_id":   kid,
            "chain":       _norm_chain(str(r[1])),
            "offer_count": int(r[2]),
            "avg_jatak":   float(r[3] or 0),
            "status":      "Pause" if is_pause else "Ophørt",
            "hk_opslag":   is_pause and (kid in active_2026),
        })

    ophoert_count = sum(1 for s in stores if s["status"] == "Ophørt")
    hk_count      = sum(1 for s in stores if s["hk_opslag"])
    ikke_aktive   = sum(1 for s in stores if s["status"] == "Pause" and not s["hk_opslag"])
    reelt_tabt    = sum(1 for s in stores if not s["hk_opslag"])  # ophørt + pause UDEN hk
    tabt_opslag   = sum(s["offer_count"] for s in stores)

    # Aggregate by chain
    chain_map: dict = defaultdict(lambda: {
        "count": 0, "total_opslag": 0, "wtd_jatak": 0.0, "ophoert": 0, "hk": 0, "ikke_aktive": 0
    })
    for s in stores:
        d = chain_map[s["chain"]]
        d["count"]        += 1
        d["total_opslag"] += s["offer_count"]
        d["wtd_jatak"]    += s["offer_count"] * s["avg_jatak"]
        if s["status"] == "Ophørt":
            d["ophoert"] += 1
        if s["hk_opslag"]:
            d["hk"] += 1
        if s["status"] == "Pause" and not s["hk_opslag"]:
            d["ikke_aktive"] += 1

    chains = [
        {
            "chain":              ch,
            "count":              d["count"],
            "total_opslag_2025":  d["total_opslag"],
            "avg_jatak_2025":     round(d["wtd_jatak"] / d["total_opslag"], 1) if d["total_opslag"] else 0.0,
            "ophoert_count":      d["ophoert"],
            "pause_count":        d["count"] - d["ophoert"],
            "hk_count":           d["hk"],
            "ikke_aktive_count":  d["ikke_aktive"],
            "reelt_tabt_count":   d["count"] - d["hk"],
        }
        for ch, d in sorted(chain_map.items(), key=lambda x: -x[1]["count"])
    ]

    return {
        "total_active_2025":  total_active_2025,
        "total_offers_2025":  total_offers_2025,
        "total_inactive":     len(stores),
        "ophoert_count":      ophoert_count,
        "pause_count":        len(stores) - ophoert_count,
        "hk_count":           hk_count,
        "ikke_aktive":        ikke_aktive,
        "reelt_tabt":         reelt_tabt,
        "new_stores_2026":    new_stores_2026,
        "tabt_opslag_2025":   tabt_opslag,
        "chains":             chains,
    }


@router.get("/stores")
def get_churn_stores(chain: str = Query(...)):
    """
    Individual inactive stores for a specific chain.
    chain = normalised chain name (e.g. 'Kvickly', "Dagli'Brugsen").
    """
    conn = get_conn()
    active_ids  = _get_active_kardex()
    active_2026 = _get_active_2026_kardex()
    name_map    = _get_store_name_map()
    safe = chain.replace("'", "''")

    rows = conn.execute(_CTE + f"""
        SELECT kardex_id, store_name, offer_count, avg_jatak, seneste_opslag
        FROM inactive
        WHERE store_name LIKE '{safe}%'
        ORDER BY offer_count DESC
    """).fetchall()

    return [
        {
            "kardex_id":      str(int(r[0])),
            "name":           name_map.get(int(r[0]), str(r[1])),
            "offer_count":    int(r[2]),
            "avg_jatak":      float(r[3] or 0),
            "seneste_opslag": str(r[4]),
            "status":         "Pause" if int(r[0]) in active_ids else "Ophørt",
            "hk_opslag":      (int(r[0]) in active_ids) and (int(r[0]) in active_2026),
        }
        for r in rows
    ]
